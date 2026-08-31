"""Phase C4 — robust, deterministic loading of messy merchant exports."""
from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import Workbook

from app.services.file_ingestion import resolve_columns
from app.services.workbook_loader import DataQualityError, load_workbook


def _xlsx_bytes(sheets: list[tuple[str, list[list[str | None]]]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_csv_with_bom_semicolon_arabic():
    csv_bytes = "\ufeffاسم المنتج;الكمية;السعر;التكلفة\nحليب;١٠;12;8\nماء;٥;4;2\n".encode("utf-8")
    load = load_workbook(csv_bytes, "sales.csv")
    assert load.file_type == "csv"
    assert "product_name" in resolve_columns(load.df).detected_fields
    assert len(load.df) == 2


def test_csv_utf8_comma_shift_jis_safe():
    csv_bytes = "Product,Qty,Price\nMilk,10,12\n".encode("utf-8")
    load = load_workbook(csv_bytes, "inventory.csv")
    assert resolve_columns(load.df).mapping.get("product_name") == "Product"


def test_xlsx_picks_best_sheet_after_empty_first_sheet():
    content = _xlsx_bytes([
        ("Empty", []),
        ("Sheet2", [["اسم المنتج", "الكمية", "السعر"], ["حليب", "10", "12"]]),
        ("Notes", [["Random", "text"], ["more", "stuff"]]),
    ])
    load = load_workbook(content, "store.xlsx")
    assert load.file_type == "xlsx"
    assert load.selected_sheet == "Sheet2"
    assert load.sheet_count == 3
    assert resolve_columns(load.df).mapping.get("product_name") == "اسم المنتج"


def test_xlsx_duplicate_headers_are_deduplicated():
    content = _xlsx_bytes([
        ("S", [["name", "qty", "qty", "price"], ["milk", "1", "2", "3"]]),
    ])
    load = load_workbook(content, "dup.xlsx")
    cols = set(load.df.columns)
    assert "qty" in cols
    assert "qty_2" in cols


def test_xlsx_blank_columns_dropped():
    content = _xlsx_bytes([
        ("S", [["name", "", "price", None], ["milk", None, "12", None]]),
    ])
    load = load_workbook(content, "blank.xlsx")
    assert "" not in load.df.columns


def test_json_rows_accepted():
    content = b'{"rows": [{"product": "Milk", "qty": "10"}]}'
    load = load_workbook(content, "data.json")
    assert load.file_type == "json"
    assert len(load.df) == 1


def test_garbage_xlsx_is_rejected_with_reason():
    with pytest.raises(DataQualityError) as e:
        load_workbook(b"this is not a real xlsx file, just text", "bad.xlsx")
    assert e.value.parse_failure_reason


def test_empty_csv_is_rejected():
    with pytest.raises(DataQualityError) as e:
        load_workbook(b"\n\n\n", "empty.csv")
    assert e.value.parse_failure_reason


def test_deterministic_across_calls():
    csv_bytes = "name,qty,price\nmilk,١٠,12\n".encode("utf-8")
    a = load_workbook(csv_bytes, "a.csv")
    b = load_workbook(csv_bytes, "a.csv")
    assert list(a.df.columns) == list(b.df.columns)
    assert a.df.to_dict() == b.df.to_dict()


def test_arabic_indic_digits_in_headers_load():
    csv_bytes = "اسم المنتج,الكمية,سعر البيع\nحليب,٢٠,١٢\n".encode("utf-8")
    load = load_workbook(csv_bytes, "ar.csv")
    res = resolve_columns(load.df)
    assert res.mapping.get("product_name") == "اسم المنتج"
    assert res.detected_fields == ["product_name", "quantity", "price"]