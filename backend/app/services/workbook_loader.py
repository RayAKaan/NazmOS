"""Deterministic, defensive loading of messy merchant exports (CSV / XLSX / XLS / JSON).

Handles: multiple sheets, empty first sheet, merged / duplicate / missing
headers, blank columns, BOM and unknown encodings, mixed numeric formats in
Arabic or English, and unexpected column order. Never guesses silently data,
and every decision is reproducible.
"""
from __future__ import annotations

import csv
import io
import json
import os
from dataclasses import dataclass
from typing import Any

import chardet
import pandas as pd
from openpyxl import load_workbook as _openpyxl_load_workbook

from app.services.file_ingestion import NORMALIZED_ALIASES, normalize_header, to_ascii_digits

MAX_SHEETS = 20
MAX_HEADER_SCAN_ROWS = 12
_ALIAS_KEYS = {alias for aliases in NORMALIZED_ALIASES.values() for alias in aliases}


class DataQualityError(ValueError):
    """Raised when a file cannot be parsed cleanly; carries a stable reason."""

    def __init__(self, parse_failure_reason: str):
        super().__init__(parse_failure_reason)
        self.parse_failure_reason = parse_failure_reason


@dataclass
class WorkbookLoad:
    df: pd.DataFrame
    file_type: str
    sheet_count: int | None = None
    selected_sheet: str | None = None
    header_row_index: int | None = None
    encoding: str | None = None
    delimiter: str | None = None

    def meta(self) -> dict[str, Any]:
        out: dict[str, Any] = {
            "file_type": self.file_type,
            "sheet_count": self.sheet_count,
            "selected_sheet": self.selected_sheet,
            "header_row_index": self.header_row_index,
            "encoding": self.encoding,
            "delimiter": self.delimiter,
        }
        return {k: v for k, v in out.items() if v is not None}


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        norm = normalize_header(h)
        if not norm:
            out.append("")
            continue
        label = h.strip()
        count = seen.get(norm, 0)
        seen[norm] = count + 1
        out.append(label if count == 0 else f"{label}_{count + 1}")
    return out


def _header_score(row: list[str]) -> int:
    score = 0
    seen = set()
    for cell in row:
        norm = normalize_header(cell)
        if not norm:
            continue
        if norm in _ALIAS_KEYS and norm not in seen:
            score += 3
            seen.add(norm)
        if norm not in _ALIAS_KEYS and norm not in seen:
            score += 1
            seen.add(norm)
    return score


def _pick_header_row(matrix: list[list[Any]]) -> tuple[int, list[str], list[list[Any]]]:
    best_row = 0
    best_score = -1
    for idx in range(min(MAX_HEADER_SCAN_ROWS, len(matrix))):
        score = _header_score(matrix[idx])
        non_empty = sum(1 for c in matrix[idx] if c not in (None, "") and normalize_header(c))
        score += non_empty * 2
        if score > best_score:
            best_score = score
            best_row = idx
    if best_score < 3:
        raise DataQualityError("No recognizable header row found. Expected columns like product name, quantity, price.")
    headers = _dedupe_headers(["" if c in (None, "") else str(c).strip() for c in matrix[best_row]])
    data = matrix[best_row + 1:]
    return best_row, headers, data


def _clean_frame(headers: list[str], data: list[list[Any]]) -> pd.DataFrame:
    width = max(len(headers), *[len(r) for r in data]) if data else len(headers)
    headers = headers + [""] * (width - len(headers))
    rows = [r + [None] * (width - len(r)) for r in data]
    df = pd.DataFrame(rows, columns=[h or f"col_{i + 1}" for i, h in enumerate(headers)])
    df = df.dropna(axis=1, how="all")
    df = df.loc[:, ~df.columns.duplicated()]
    return df


def _load_csv(content: bytes) -> tuple[pd.DataFrame, dict[str, Any]]:
    sample = content[:8192]
    encoding = None
    for candidate in ("utf-8-sig", "utf-8"):
        try:
            sample.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if encoding is None:
        guess = chardet.detect(sample)
        encoding = guess.get("encoding") or "utf-8"
        for candidate in (encoding, "cp1256", "latin-1"):
            try:
                sample.decode(candidate)
                encoding = candidate
                break
            except (UnicodeDecodeError, LookupError):
                continue

    text = content.decode(encoding, errors="ignore")
    sample_text = text[:4096]
    delimiter = ","
    try:
        sniffed = csv.Sniffer().sniff(sample_text, delimiters=",;\t|")
        delimiter = sniffed.delimiter
    except csv.Error:
        lines = sample_text.splitlines()
        if not lines or not lines[0]:
            raise DataQualityError("CSV file is empty.")
        for candidate in (",", ";", "\t", "|"):
            rows = list(csv.reader(io.StringIO(lines[0]), delimiter=candidate))
            if rows and len(rows[0]) >= 2:
                delimiter = candidate
                break

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    matrix = [[None if cell in ("", "nan", "null") else cell for cell in row] for row in reader]
    matrix = [row for row in matrix if any(c for c in row)]
    if not matrix:
        raise DataQualityError("CSV file is empty.")
    header_idx, headers, data = _pick_header_row(matrix)
    df = _clean_frame(headers, data)
    return df, {"encoding": encoding, "delimiter": delimiter, "header_row_index": header_idx, "sheet_count": 1, "selected_sheet": "csv"}


def _load_xlsx(content: bytes) -> tuple[pd.DataFrame, dict[str, Any]]:
    try:
        wb = _openpyxl_load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise DataQualityError(f"Could not open XLSX workbook: {exc}")
    sheet_count = min(len(wb.sheetnames), MAX_SHEETS)
    best: tuple[float, str, int, list[list[Any]]] | None = None
    for sheet_name in wb.sheetnames[:MAX_SHEETS]:
        ws = wb[sheet_name]
        matrix = []
        for row in ws.iter_rows(values_only=True):
            matrix.append([None if v is None else _cell_to_str(v) for v in row])
        matrix = [row for row in matrix if any(c for c in row)]
        if not matrix:
            continue
        try:
            header_idx, headers, data = _pick_header_row(matrix)
        except DataQualityError:
            continue
        df = _clean_frame(headers, data)
        alias_hits = sum(1 for h in df.columns if normalize_header(h) in _ALIAS_KEYS)
        score = alias_hits + min(len(df), 100) / 100.0
        if best is None or score > best[0]:
            best = (score, sheet_name, header_idx, matrix)
    if best is None:
        raise DataQualityError("XLSX file contains no sheet with a recognizable header row.")
    _, sheet_name, header_idx, matrix = best
    header_idx, headers, data = _pick_header_row(matrix)
    df = _clean_frame(headers, data)
    return df, {"header_row_index": header_idx, "sheet_count": sheet_count, "selected_sheet": sheet_name}


def _load_xls_or_other(content: bytes, filename: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".xls":
        try:
            df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False)
        except Exception as exc:
            raise DataQualityError(f"Could not read XLS workbook: {exc}")
        matrix = [[None if v in ("", "nan", "null") else v for v in row] for row in df.itertuples(index=False)]
        if matrix:
            header_idx, headers, data = _pick_header_row(matrix)
            df = _clean_frame(headers, data)
        return df, {"header_row_index": 0, "sheet_count": 1, "selected_sheet": "sheet1"}
    raise DataQualityError(f"Unsupported file type: {ext}")


def _load_json(content: bytes) -> tuple[pd.DataFrame, dict[str, Any]]:
    try:
        data = json.loads(content.decode("utf-8", errors="ignore"))
    except Exception as exc:
        raise DataQualityError(f"Invalid JSON file: {exc}")
    if isinstance(data, dict):
        data = data.get("rows", data.get("data", []))
    if not isinstance(data, list) or not data:
        raise DataQualityError("JSON file contains no rows.")
    df = _clean_frame(list(data[0].keys()), [[r.get(k) for k in data[0].keys()] for r in data])
    return df, {"header_row_index": 0, "sheet_count": 1, "selected_sheet": "json"}


def _cell_to_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if hasattr(value, "isoformat"):
        return str(value.date()) if hasattr(value, "date") else str(value)
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
        return repr(value)
    return str(value).strip()


def load_workbook(content: bytes, filename: str) -> WorkbookLoad:
    """Load any supported export into a clean frame with unique headers."""
    ext = os.path.splitext(filename)[1].lower()
    if ext in {".xlsx", ".xlsm"}:
        df, extra = _load_xlsx(content)
        file_type = "xlsx"
    elif ext == ".csv":
        df, extra = _load_csv(content)
        file_type = "csv"
    elif ext == ".xls":
        df, extra = _load_xls_or_other(content, filename)
        file_type = "xls"
    elif ext == ".json":
        df, extra = _load_json(content)
        file_type = "json"
    else:
        raise DataQualityError(f"Unsupported file extension: {ext or 'none'}")

    df.columns = [str(c).strip() for c in df.columns]
    return WorkbookLoad(
        df=df,
        file_type=file_type,
        sheet_count=extra.get("sheet_count"),
        selected_sheet=extra.get("selected_sheet"),
        header_row_index=extra.get("header_row_index"),
        encoding=extra.get("encoding"),
        delimiter=extra.get("delimiter"),
    )


def normalize_frame_numbers(df: pd.DataFrame) -> pd.DataFrame:
    """Coerce every string cell that looks numeric (Arabic/English) to Decimal-compatible text.

    Returns a string-frame with normalized numerals so downstream parsing is
    deterministic regardless of locale formatting.
    """
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(lambda v: to_ascii_digits(v) if isinstance(v, (str, int, float)) else v)
    return out